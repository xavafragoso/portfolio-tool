# %% [markdown]
# # export — Generación de Excel, PDF y JSON institucional
#
# Módulos O y P del notebook legacy + exportación JSON para el dashboard. Contiene:
# - `_t()`  : sanitizador de texto a latin-1 para fpdf2 (fix de encoding ya
#             resuelto — acentos y guiones largos). **No modificar.**
# - `PortfolioPDF` : clase FPDF con header/footer/tablas/KPIs del legacy.
# - `exportar_excel()` / `exportar_pdf()` : extraídas del callback monolítico.
# - `exportar_json()` : [Fase 2] arma `outputs/data.json` con la base analítica
#             (derivada de `R`) + las 3 secciones nuevas (fundamental, señales,
#             noticias). Ese archivo alimentará `dashboard/index.html` en Fase 3.
#
# Adaptaciones respecto a Colab:
# - Se elimina `google.colab.files.download()`; los archivos se escriben a
#   la carpeta `outputs/` en disco.
# - `DataFrame.applymap` → `DataFrame.map` (applymap fue removido en pandas 3.0).
#
# Las funciones de exportación reciben un dict de resultados `R` armado por el
# orquestador (`notebooks/analisis.py`), para no reproducir aquí el pipeline.

# %%
import os
import json
from datetime import datetime

import numpy as np
import pandas as pd

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

from stress import PERIODOS_STRESS

OUTPUT_DIR = 'outputs'


# %% [markdown]
# ## Utilidad: sanitizador de texto para fpdf2 (fix de encoding — NO modificar)

# %%
def _t(texto):
    """Sanitiza texto a latin-1 para fpdf2."""
    mapa = {'—': '--', '–': '-', '’': "'", '“': '"', '”': '"',
            '‘': "'", 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'á': 'a', 'ñ': 'n', 'Á': 'A', 'É': 'E', 'Í': 'I',
            'Ó': 'O', 'Ú': 'U', 'Ñ': 'N', 'ü': 'u', 'ä': 'a',
            'μ': 'u', 'σ': 's', '≥': '>=', '≤': '<='}
    for k, v in mapa.items(): texto = str(texto).replace(k, v)
    return ''.join(c if ord(c) < 256 else '?' for c in texto)


# %% [markdown]
# ## Clase PortfolioPDF (FPDF institucional)

# %%
class PortfolioPDF(FPDF):
    C_BG = (13, 17, 23); C_PANEL = (22, 27, 34); C_ACCENT = (88, 166, 255)
    C_GOLD = (255, 215, 0); C_TEXT = (201, 209, 217); C_MUTED = (139, 148, 158)

    def header(self):
        self.set_fill_color(*self.C_BG); self.rect(0, 0, 210, 297, 'F')
        self.set_fill_color(*self.C_PANEL); self.rect(0, 0, 210, 16, 'F')
        self.set_fill_color(*self.C_ACCENT); self.rect(0, 16, 210, 0.8, 'F')
        self.set_font('Helvetica', 'B', 8); self.set_text_color(*self.C_ACCENT)
        self.set_y(4); self.cell(0, 6, _t('PORTFOLIO ANALYSIS v3 -- INSTITUTIONAL'), align='L')
        self.set_text_color(*self.C_MUTED); self.set_font('Helvetica', '', 7)
        self.set_y(4); self.cell(0, 6, _t(datetime.today().strftime('%d/%m/%Y')), align='R')
        self.ln(14)

    def footer(self):
        self.set_fill_color(*self.C_PANEL); self.rect(0, 285, 210, 12, 'F')
        self.set_fill_color(*self.C_ACCENT); self.rect(0, 285, 210, 0.5, 'F')
        self.set_y(-10); self.set_font('Helvetica', '', 7)
        self.set_text_color(*self.C_MUTED)
        self.cell(0, 5, _t(f'Modulo 6 -- Finanzas Cuantitativas | Pag {self.page_no()}'), align='C')

    def sec(self, txt):
        self.set_fill_color(*self.C_ACCENT); self.rect(self.get_x(), self.get_y(), 3, 7, 'F')
        self.set_x(self.get_x() + 6); self.set_font('Helvetica', 'B', 13)
        self.set_text_color(*self.C_ACCENT); self.cell(0, 7, _t(txt), ln=True); self.ln(4)

    def tabla(self, df, dec=4):
        cols = [str(df.index.name or 'Index')] + [str(c) for c in df.columns]
        aw = 190 / len(cols)
        self.set_fill_color(*self.C_PANEL); self.set_font('Helvetica', 'B', 8)
        self.set_text_color(*self.C_ACCENT)
        for c in cols: self.cell(aw, 7, _t(c[:16]), border=1, align='C', fill=True)
        self.ln()
        self.set_font('Helvetica', '', 7)
        for i, (idx, row) in enumerate(df.iterrows()):
            self.set_fill_color(*(self.C_BG if i % 2 == 0 else self.C_PANEL))
            self.set_text_color(*self.C_TEXT)
            self.cell(aw, 6, _t(str(idx)[:16]), border=1, align='L', fill=True)
            for val in row:
                try: txt2 = f'{float(val):.{dec}f}'
                except: txt2 = _t(str(val)[:14])
                self.cell(aw, 6, txt2, border=1, align='R', fill=True)
            self.ln()
        self.ln(5)

    def kpi_row(self, items):
        n = len(items); aw = (190 - 4 * (n - 1)) / n; y0, x0 = self.get_y(), self.get_x()
        for i, (val, lbl, sub) in enumerate(items):
            xi = x0 + i * (aw + 4); self.set_xy(xi, y0)
            self.set_fill_color(*self.C_PANEL); self.rect(xi, y0, aw, 22, 'F')
            self.set_font('Helvetica', '', 6); self.set_text_color(*self.C_MUTED)
            self.set_xy(xi, y0 + 1); self.cell(aw, 6, _t(lbl.upper()), align='C')
            self.set_font('Helvetica', 'B', 13); self.set_text_color(*self.C_ACCENT)
            self.set_xy(xi, y0 + 7); self.cell(aw, 8, _t(val), align='C')
            self.set_font('Helvetica', '', 6); self.set_text_color(*self.C_MUTED)
            self.set_xy(xi, y0 + 16); self.cell(aw, 5, _t(sub), align='C')
        self.set_y(y0 + 28)

    def img_safe(self, path, **kw):
        if os.path.exists(path): self.image(path, **kw)


# %% [markdown]
# ## Exportación a Excel
#
# Replica las hojas y el estilo del legacy. Recibe el dict de resultados `R`.

# %%
def exportar_excel(R, ruta=None):
    """Genera el libro Excel institucional con todas las hojas del legacy.

    `R` es el dict de resultados del pipeline (ver `notebooks/analisis.py`).
    Retorna la ruta del archivo escrito en `outputs/`.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if ruta is None:
        ruta = os.path.join(OUTPUT_DIR, 'portfolio_v3.xlsx')

    TICKERS = R['TICKERS']; AÑOS_MC = R['AÑOS_MC']
    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
        R['precios'].round(4).to_excel(writer, sheet_name='Precios')
        R['rend'].round(6).to_excel(writer, sheet_name='Rendimientos')
        R['stats'].round(6).to_excel(writer, sheet_name='Estadisticas')
        R['corr'].round(4).to_excel(writer, sheet_name='Correlacion')
        R['cov_lw'].round(8).to_excel(writer, sheet_name='Cov_LedoitWolf')
        R['df_sim'].round(6).to_excel(writer, sheet_name='Frontera', index=False)
        R['df_riesgo'].round(6).to_excel(writer, sheet_name='VaR_CVaR')
        R['df_stress'].round(4).to_excel(writer, sheet_name='Stress_Testing')
        R['df_reb_pms'].round(4).to_excel(writer, sheet_name='Rebalanceo_PMS')
        R['df_reb_pmv'].round(4).to_excel(writer, sheet_name='Rebalanceo_PMV')
        R['reg_stats'].round(4).to_excel(writer, sheet_name='Regimenes')
        for p, nm in [(R['pmv'], 'PMV'), (R['pms'], 'PMS'),
                      (R['prp'], 'RiskParity'), (R['pbl'], 'BlackLitterman')]:
            dp = pd.DataFrame({'Ticker': TICKERS, 'Peso': p['pesos']})
            dp.to_excel(writer, sheet_name=nm, index=False)
        # MC summary
        mc_rows = R['mc_rows']
        pd.DataFrame(mc_rows).to_excel(writer, sheet_name='MonteCarlo', index=False)
        if R.get('betas_pmv_df') is not None:
            R['betas_pmv_df'].round(4).to_excel(writer, sheet_name='FF_PMV')
            R['betas_pms_df'].round(4).to_excel(writer, sheet_name='FF_PMS')

    # Estilo Excel
    wb = openpyxl.load_workbook(ruta)
    for ws in wb.worksheets:
        fh = PatternFill('solid', fgColor='1c2333')
        fa = PatternFill('solid', fgColor='0d1117')
        fn = PatternFill('solid', fgColor='161b22')
        for ri, row in enumerate(ws.iter_rows()):
            for cell in row:
                cell.border = Border(left=Side(style='thin', color='30363d'),
                                     right=Side(style='thin', color='30363d'),
                                     top=Side(style='thin', color='30363d'),
                                     bottom=Side(style='thin', color='30363d'))
                if ri == 0:
                    cell.fill = fh; cell.font = Font(bold=True, color='58a6ff', name='Calibri', size=10)
                else:
                    cell.fill = fa if ri % 2 == 0 else fn
                    cell.font = Font(color='c9d1d9', name='Calibri', size=9)
                cell.alignment = Alignment(horizontal='right', vertical='center')
        for col in ws.columns:
            mx = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 28)
    wb.save(ruta)
    print(f'✅ Excel: {ruta} ({os.path.getsize(ruta) // 1024} KB)')
    return ruta, wb.sheetnames


# %% [markdown]
# ## Exportación a PDF institucional
#
# Replica el reporte multi-página del legacy. Opcionalmente embebe las
# gráficas Plotly (`R['figs']`) como PNG vía kaleido si está disponible.

# %%
def exportar_pdf(R, ruta=None, img_dir=None):
    """Genera el reporte PDF institucional. Retorna la ruta del archivo.

    Si kaleido está instalado, embebe las gráficas pasadas en `R['figs']`.
    Si no, genera el PDF sin gráficas embebidas (igual que el fallback legacy).
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if ruta is None:
        ruta = os.path.join(OUTPUT_DIR, 'portfolio_report_v3.pdf')
    if img_dir is None:
        img_dir = os.path.join(OUTPUT_DIR, 'pdf_imgs_v3')
    os.makedirs(img_dir, exist_ok=True)

    TICKERS = R['TICKERS']; N = R['N']; shrinkage = R['shrinkage']
    pmv, pms, prp, pbl = R['pmv'], R['pms'], R['prp'], R['pbl']
    stats = R['stats']; RF_A = R['RF_A']; NAV = R['NAV']
    FECHA_INI = R['FECHA_INI']; FECHA_FIN = R['FECHA_FIN']
    N_SIMS = R['N_SIMS']; AÑOS_MC = R['AÑOS_MC']
    riesgo_pms = R['riesgo_pms']; df_riesgo = R['df_riesgo']; df_stress = R['df_stress']
    max_dd_pmv, max_dd_pms = R['max_dd_pmv'], R['max_dd_pms']
    calmar_pmv, calmar_pms = R['calmar_pmv'], R['calmar_pms']
    rec_pmv, rec_pms = R['rec_pmv'], R['rec_pms']
    mc_rows = R['mc_rows']; df_reb_pms = R['df_reb_pms']; inds = R['inds']
    betas_pmv_df = R.get('betas_pmv_df'); betas_pms_df = R.get('betas_pms_df')
    PESO_MIN, PESO_MAX = R['PESO_MIN'], R['PESO_MAX']
    BL_TAU, BL_CONF = R['BL_TAU'], R['BL_CONF']
    figs = R.get('figs', {})

    # ── Renderizado de imágenes PNG (kaleido) ────────────────────────────────
    IMGS_OK = False
    try:
        import kaleido  # noqa
        if 'frontera' in figs: figs['frontera'].write_image(f'{img_dir}/frontera.png', width=900, height=480, scale=2)
        if 'corr' in figs: figs['corr'].write_image(f'{img_dir}/corr.png', width=900, height=max(400, 52 * N), scale=2)
        if 'stress' in figs: figs['stress'].write_image(f'{img_dir}/stress.png', width=900, height=380, scale=2)
        if 'dd' in figs: figs['dd'].write_image(f'{img_dir}/dd.png', width=900, height=430, scale=2)
        if 'rolling' in figs: figs['rolling'].write_image(f'{img_dir}/rolling.png', width=900, height=550, scale=2)
        if 'ff' in figs: figs['ff'].write_image(f'{img_dir}/ff.png', width=900, height=400, scale=2)
        if 'mc' in figs: figs['mc'].write_image(f'{img_dir}/mc.png', width=1200, height=380 * len(AÑOS_MC), scale=1.5)
        IMGS_OK = True; print('  Imagenes PNG OK')
    except Exception as e:
        print(f'  kaleido: {e} -- PDF sin graficas embebidas')

    pdf = PortfolioPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.set_margins(10, 20, 10)

    # Portada
    pdf.add_page()
    pdf.set_fill_color(*pdf.C_BG); pdf.rect(0, 0, 210, 297, 'F')
    pdf.set_fill_color(*pdf.C_ACCENT); pdf.rect(0, 85, 210, 2, 'F'); pdf.rect(0, 205, 210, 2, 'F')
    pdf.set_y(100); pdf.set_font('Helvetica', 'B', 28)
    pdf.set_text_color(*pdf.C_ACCENT); pdf.cell(0, 14, 'PORTFOLIO ANALYSIS', align='C', ln=True)
    pdf.set_text_color(255, 255, 255); pdf.cell(0, 14, 'INSTITUTIONAL v3', align='C', ln=True)
    pdf.ln(6); pdf.set_font('Helvetica', '', 11); pdf.set_text_color(*pdf.C_MUTED)
    pdf.cell(0, 6, _t('Markowitz + Black-Litterman + Risk Parity + CVaR + Stress + FF + MC'), align='C', ln=True)
    pdf.ln(8); pdf.set_font('Helvetica', 'B', 9); pdf.set_text_color(*pdf.C_GOLD)
    pdf.cell(0, 5, _t(f'Activos: {" | ".join(TICKERS)}'), align='C', ln=True)
    pdf.ln(3); pdf.set_font('Helvetica', '', 8); pdf.set_text_color(*pdf.C_MUTED)
    pdf.cell(0, 5, _t(f'NAV: ${NAV:,.0f} | RF: {RF_A:.2%} | LW Shrinkage: {shrinkage:.4f}'), align='C', ln=True)
    pdf.cell(0, 5, _t(f'{FECHA_INI} a {FECHA_FIN} | {datetime.today().strftime("%d/%m/%Y %H:%M")}'), align='C', ln=True)

    # Resumen ejecutivo
    pdf.add_page(); pdf.sec('1. Resumen Ejecutivo')
    pdf.set_font('Helvetica', '', 9); pdf.set_text_color(*pdf.C_TEXT)
    pdf.multi_cell(0, 5, _t(f'Analisis cuantitativo institucional de {N} activos. '
        f'Covarianza estimada con Ledoit-Wolf (shrinkage={shrinkage:.4f}). '
        f'Cuatro portafolios: PMV, PMS (Markowitz), Risk Parity y Black-Litterman. '
        f'Riesgo: VaR parametrico, historico y Monte Carlo al 95/99%. CVaR/Expected Shortfall. '
        f'Stress testing en {len(PERIODOS_STRESS)} periodos historicos. '
        f'Rolling analysis 1Y/2Y/3Y. Factores Fama-French proxy. Regimenes de mercado K-Means. '
        f'Monte Carlo GBM con {N_SIMS:,} simulaciones. Trade list con costo estimado.'))
    pdf.ln(8)
    pdf.set_font('Helvetica', 'B', 10); pdf.set_text_color(*pdf.C_ACCENT)
    pdf.cell(0, 6, _t('KPIs — Portafolio Max Sharpe (PMS)'), ln=True); pdf.ln(3)  # FIX (port): faltaba _t() en el legacy
    pdf.kpi_row([
        (f"{pms['rendimiento']:.2%}", 'Rend. Anual', 'PMS esperado'),
        (f"{pms['riesgo']:.2%}", 'Sigma Anual', 'PMS anualizado'),
        (f"{pms['sharpe']:.3f}", 'Ratio Sharpe', 'PMS optimizado'),
        (f"{riesgo_pms['CVaR 99% (ES)']:.2%}", 'CVaR 99%', 'Expected Shortfall'),
        (f"{max_dd_pms:.2%}", 'Max Drawdown', 'Peor caida historica'),
        (f"{calmar_pms:.2f}", 'Calmar Ratio', 'Rend/MaxDD'),
    ])

    # Estadísticos
    pdf.sec('2. Estadisticos por Activo')
    pdf.tabla(stats[['Rend_Anual', 'Std_Anual', 'Sharpe', 'Skewness', 'Kurtosis']].round(4))

    # Portafolios
    pdf.add_page(); pdf.sec('3. Portafolios Optimizados')
    for p, lbl in [(pmv, 'Min Varianza PMV'), (pms, 'Max Sharpe PMS'),
                   (prp, 'Risk Parity'), (pbl, 'Black-Litterman')]:
        pdf.set_font('Helvetica', 'B', 9); pdf.set_text_color(*pdf.C_GOLD)
        pdf.cell(0, 6, lbl, ln=True)
        pdf.set_font('Helvetica', '', 8); pdf.set_text_color(*pdf.C_TEXT)
        pdf.cell(0, 5, _t(f'Rend:{p["rendimiento"]:.2%} sigma:{p["riesgo"]:.2%} Sharpe:{p["sharpe"]:.3f}'), ln=True)
        pdf.ln(1)
        dp = pd.DataFrame({'Ticker': TICKERS, 'Peso': [f'{w:.2%}' for w in p['pesos']]}).set_index('Ticker')
        pdf.tabla(dp); pdf.ln(3)

    # Imagenes
    for f, tit in [('frontera.png', '4. Frontera Eficiente + CML'),
                   ('corr.png', '5. Matriz de Correlacion (Ledoit-Wolf)'),
                   ('dd.png', '6. Drawdown — Underwater Chart'),
                   ('stress.png', '7. Stress Testing Historico'),
                   ('rolling.png', '8. Rolling Sharpe 1Y/2Y/3Y'),
                   ('ff.png', '9. Factores Fama-French'),
                   ('mc.png', '10. Monte Carlo GBM')]:
        if IMGS_OK and os.path.exists(f'{img_dir}/{f}'):
            pdf.add_page(); pdf.sec(tit)
            pdf.img_safe(f'{img_dir}/{f}', x=10, w=190)

    # Riesgo
    pdf.add_page(); pdf.sec('11. Riesgo — VaR / CVaR / Expected Shortfall')
    pdf.tabla(df_riesgo.map(lambda x: f'{x:.4f}' if isinstance(x, float) else str(x)))

    # Stress
    pdf.add_page(); pdf.sec('12. Stress Testing — Detalle')
    pdf.tabla(df_stress.reset_index().set_index('Periodo' if 'Periodo' in df_stress.reset_index().columns else df_stress.reset_index().columns[0]).round(4))

    # Drawdown
    pdf.add_page(); pdf.sec('13. Drawdown Analysis')
    dd_data = pd.DataFrame({'Max DD': [max_dd_pmv, max_dd_pms], 'Calmar': [calmar_pmv, calmar_pms],
                            'Rec. Dias': [rec_pmv or 'No rec.', rec_pms or 'No rec.']},
                           index=['PMV', 'PMS'])
    pdf.tabla(dd_data)

    # MC
    pdf.add_page(); pdf.sec('14. Monte Carlo GBM — Resumen')
    mc_tab = pd.DataFrame(mc_rows).set_index(['Port', 'Anos'])
    pdf.tabla(mc_tab.round(4))

    # Rebalanceo
    pdf.add_page(); pdf.sec('15. Trade List — Rebalanceo hacia PMS')
    reb_show = df_reb_pms[['Peso Actual', 'Peso Objetivo', 'Gap', 'Direccion' if 'Direccion' in df_reb_pms.columns else 'Dirección', 'USD Trade', 'Comision Est.' if 'Comision Est.' in df_reb_pms.columns else 'Comisión Est.']].copy()
    reb_show.columns = [_t(c) for c in reb_show.columns]
    pdf.tabla(reb_show.round(4))
    pdf.set_font('Helvetica', '', 8); pdf.set_text_color(*pdf.C_MUTED)
    pdf.cell(0, 5, _t(f'Turnover: ${df_reb_pms["USD Trade"].sum():,.0f} ({df_reb_pms["USD Trade"].sum()/NAV:.1%} del NAV) | Costo est.: ${df_reb_pms["Comision Est."].sum() if "Comision Est." in df_reb_pms.columns else df_reb_pms["Comisión Est."].sum():,.0f}'), ln=True)

    # Señales técnicas
    pdf.add_page(); pdf.sec('16. Senales Tecnicas (90 dias)')
    for tk, ind2 in inds.items():
        pdf.set_font('Helvetica', 'B', 9); pdf.set_text_color(*pdf.C_GOLD)
        pdf.cell(0, 6, tk, ln=True)
        pdf.set_font('Helvetica', '', 8); pdf.set_text_color(*pdf.C_TEXT)
        for s in ind2['señales']: pdf.cell(0, 5, _t(f'  {s}'), ln=True)
        pdf.ln(2)

    # Notas metodológicas
    pdf.add_page(); pdf.sec('17. Notas Metodologicas')
    notas = [
        ('Rendimientos', 'Logaritmicos ln(Pt/Pt-1). Anualizacion: mu*F, sigma*sqrt(F).'),
        ('Ledoit-Wolf', f'Shrinkage={shrinkage:.4f}. Reduce error estimacion matriz covarianza.'),
        ('Markowitz', f'SLSQP. Pesos [{PESO_MIN:.0%},{PESO_MAX:.0%}]. Sin cortos.'),
        ('Black-Litterman', f'tau={BL_TAU}, confianza={BL_CONF:.0f}%. Equilibrio+views PM.'),
        ('Risk Parity', 'Equal Risk Contribution. Minimiza dispersion contribuciones de riesgo.'),
        ('CVaR/ES', 'Expected Shortfall: perdida esperada dado que supera el VaR.'),
        ('Stress', '7 periodos historicos de crisis. Retorno total y max DD real.'),
        ('Drawdown', 'Calmar=Rend_anual/|MaxDD|. Tiempo recuperacion desde trough.'),
        ('Fama-French', 'Proxy ETF: Mkt=SPY, SMB=IWM-SPY, HML=IVE-IWF, MOM=MTUM-USMV.'),
        ('Regimenes', 'K-Means sobre ret/vol rolling 21d. 3 estados: Bull/Lateral/Bear.'),
        ('Monte Carlo', f'GBM, {N_SIMS:,} sims. S_t=S_0*exp((mu-0.5*sig^2)*t+sig*sqrt(t)*Z).'),
        ('Rebalanceo', 'Delta acciones = (NAV*w_obj/precio) - posicion_actual. Costo=delta*precio*comision.'),
    ]
    pdf.set_font('Helvetica', '', 8)
    for con, desc in notas:
        pdf.set_text_color(*pdf.C_ACCENT); pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(38, 5, _t(con), border='B')
        pdf.set_text_color(*pdf.C_TEXT); pdf.set_font('Helvetica', '', 8)
        pdf.multi_cell(0, 5, _t(desc), border='B'); pdf.ln(1)

    pdf.output(ruta)
    print(f'✅ PDF: {ruta} ({os.path.getsize(ruta) // 1024} KB)')
    return ruta


# %% [markdown]
# ## [Fase 2] Exportación a JSON para el dashboard
#
# `export.py` no generaba JSON antes; esta función crea `outputs/data.json` con
# una **base analítica** derivada del dict de resultados `R` (estadísticos,
# portafolios, riesgo, stress, Monte Carlo, rebalanceo, regímenes) y le agrega
# **al final** las 3 secciones nuevas de Fase 2: `fundamental`, `senales`,
# `noticias`. Ese archivo lo consumirá `dashboard/index.html` en Fase 3.

# %%
def _a_serializable(obj):
    """Convierte estructuras con tipos numpy/pandas a algo serializable a JSON.

    DataFrames → lista de registros; Series → dict; escalares numpy → nativos;
    NaN/Inf → None; Timestamps → ISO. Recorre dicts y listas de forma recursiva.
    """
    if isinstance(obj, dict):
        return {str(k): _a_serializable(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_a_serializable(v) for v in obj]
    if isinstance(obj, pd.DataFrame):
        return _a_serializable(obj.reset_index().to_dict('records'))
    if isinstance(obj, pd.Series):
        return _a_serializable(obj.to_dict())
    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()
    if isinstance(obj, np.generic):
        obj = obj.item()
    if isinstance(obj, np.ndarray):
        return _a_serializable(obj.tolist())
    if isinstance(obj, float):
        return obj if np.isfinite(obj) else None
    return obj


def exportar_json(R, fundamental=None, senales=None, noticias=None, ruta=None):
    """Escribe `outputs/data.json` con la base analítica + secciones Fase 2.

    Parámetros
    ----------
    R          : dict de resultados del pipeline (ver `notebooks/analisis.py`).
    fundamental: salida de `fundamental.obtener_fundamentales()`.
    senales    : salida de `signals.generar_senales()`.
    noticias   : salida de `news.analizar_noticias()`.
    ruta       : destino; por defecto `outputs/data.json`.

    Retorna la ruta del archivo escrito.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if ruta is None:
        ruta = os.path.join(OUTPUT_DIR, 'data.json')

    tickers = R.get('TICKERS', [])

    # Portafolios optimizados (pesos por ticker)
    portafolios = {}
    for key, nm in [('pmv', 'PMV'), ('pms', 'PMS'),
                    ('prp', 'RiskParity'), ('pbl', 'BlackLitterman')]:
        p = R.get(key)
        if p is None:
            continue
        portafolios[nm] = {
            'rendimiento': p.get('rendimiento'),
            'riesgo':      p.get('riesgo'),
            'sharpe':      p.get('sharpe'),
            'pesos':       {t: w for t, w in zip(tickers, p.get('pesos', []))},
        }

    # ── Secciones que ya formaban parte del análisis (base) ──────────────────
    data = {
        'meta': {
            'tickers':   tickers,
            'n_activos': R.get('N'),
            'nav':       R.get('NAV'),
            'rf_anual':  R.get('RF_A'),
            'shrinkage': R.get('shrinkage'),
            'fecha_ini': R.get('FECHA_INI'),
            'fecha_fin': R.get('FECHA_FIN'),
            'generado':  datetime.today().isoformat(timespec='seconds'),
        },
        'estadisticos':   R.get('stats'),
        'portafolios':    portafolios,
        'riesgo':         R.get('df_riesgo'),
        'stress':         R.get('df_stress'),
        'montecarlo':     R.get('mc_rows'),
        'rebalanceo_pms': R.get('df_reb_pms'),
        'regimenes':      R.get('reg_stats'),
    }

    # ── Secciones nuevas de Fase 2 (agregadas al final) ──────────────────────
    data['fundamental'] = fundamental
    data['senales']     = senales
    data['noticias']    = noticias

    limpio = _a_serializable(data)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(limpio, f, ensure_ascii=False, indent=2, allow_nan=False)
    print(f'✅ JSON: {ruta} ({os.path.getsize(ruta) // 1024} KB)')
    return ruta
